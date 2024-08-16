# Price is wrong on all cells, need to decrease by 10%
# Add charts
import openpyxl as xl
from openpyxl.chart import PieChart, Reference

def workbook_selection(fileName):
    wbo = xl.load_workbook('sales.xlsx')
    sheet = wbo['Sheet1']  # Sheet name is case-sensitive

    # Iterate through rows starting from row 2 to avoid the header row
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)  # Get the cell in column 3 (assumed to be the price column)
        changed_price = cell.value * 0.9  # Decrease the price by 10%
        changed_price_cell = sheet.cell(row, 4)  # Place the new price in column 4
        changed_price_cell.value = changed_price  # Set the new price value

    # Adding a pie chart

    # Select values from column 4 (new prices) for all rows
    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)

    pie_chart = PieChart()
    pie_chart.add_data(values)  # Add the selected data to the pie chart
    sheet.add_chart(pie_chart, 'a6')  # Place the pie chart at cell A6

    # Save the modified workbook
    wbo.save(fileName)

# Run the function to process the Excel file
workbook_selection("sales.xlsx")
